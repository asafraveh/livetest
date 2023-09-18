import tkinter as tk
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from openpyxl import load_workbook
from tkcalendar import Calendar
import json
import threading
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

def open_website_and_login(website_url, username_value, password_value):
    driver = webdriver.Chrome()
    driver.get(website_url)
    wait = WebDriverWait(driver, 10)
    username_field = wait.until(EC.element_to_be_clickable((By.ID, "username")))
    username_field.click()
    username_field.send_keys(username_value)
    password_field = wait.until(EC.element_to_be_clickable((By.ID, "password")))
    password_field.click()
    password_field.send_keys(password_value)
    login_button = wait.until(EC.element_to_be_clickable((By.ID, "kc-login")))
    login_button.click()
    input("Press Enter to close the browser...")
    driver.quit()

def open_website_thread(website_url, username_value, password_value):
    t = threading.Thread(target=open_website_and_login, args=(website_url, username_value, password_value))
    t.start()

def clear_fields():
    net_entry.delete(0, tk.END)
    version_entry.delete(0, tk.END)
    total_event_entry.delete(0, tk.END)
    false_event_entry.delete(0, tk.END)
    version_bugs_entry.delete(0, tk.END)
    unique_id_entry.delete(0, tk.END)
    calendar.set_date("")  # Clear the calendar date

root = tk.Tk()
root.title("Website Automation")

# Create a frame for the website buttons
website_frame = ttk.LabelFrame(root, text="Websites")
website_frame.pack(padx=10, pady=10, fill="both", expand="yes")

button1 = tk.Button(website_frame, text="CEMEX", command=lambda: open_website_thread(
    "https://cemex-manager-app.ception.live/",
    "ceptionuser",
    "ceptioncemex?"
))
button2 = tk.Button(website_frame, text="Shafir", command=lambda: open_website_thread(
    "https://shapir-manager-app.ception.live/",
    "ceptionuser",
    "ceptionshapir!"
))
button3 = tk.Button(website_frame, text="Heidelberg", command=lambda: open_website_thread(
    "https://heidelberg-manager-app.ception.live/",
    "ceptionuser",
    "ceptionheidelberg%"
))

button1.grid(row=0, column=0, padx=5, pady=5)
button2.grid(row=0, column=1, padx=5, pady=5)
button3.grid(row=0, column=2, padx=5, pady=5)

# Create a frame for data entry
data_frame = ttk.LabelFrame(root, text="Data Entry")
data_frame.pack(padx=10, pady=10, fill="both", expand="yes")

site_options = {
    "Shafir": ["roller", "Etziyona"],
    "cemex": ["adhalom", "Gdansk", "Gdania", "golani", "mevocarmel", "modiim", "yvnehe"],
    "hidenberg": ["London1627",
         "London1656"]
}

def update_site_options(event):
    selected_customer = customer_combo.get()
    site_combo["values"] = site_options.get(selected_customer, [])

def update_net_version(event):
    selected_site = site_combo.get()
    site_data = site_config.get(selected_site, {})
    net_entry.delete(0, tk.END)
    net_entry.insert(0, site_data.get("net", ""))
    version_entry.delete(0, tk.END)
    version_entry.insert(0, site_data.get("version", ""))

def add_data():
    customer = customer_combo.get()
    selected_sites = site_combo.get()
    selected_sites = selected_sites if isinstance(selected_sites, list) else [selected_sites]
    date = calendar.get_date()
    net = net_entry.get()
    version = version_entry.get()
    total_event = total_event_entry.get()
    false_event = false_event_entry.get()
    version_bugs = version_bugs_entry.get()
    unique_id = unique_id_entry.get()

    try:
        workbook = load_workbook("O:\\QA\\live.xlsx")
        sheet = workbook[customer]

        for site in selected_sites:
            row_number = sheet.max_row + 1

            sheet.cell(row=row_number, column=1, value=date)
            sheet.cell(row=row_number, column=2, value=site)
            sheet.cell(row=row_number, column=3, value=net)
            sheet.cell(row=row_number, column=4, value=version)
            sheet.cell(row=row_number, column=5, value=total_event)
            sheet.cell(row=row_number, column=6, value=false_event)
            sheet.cell(row=row_number, column=7, value=version_bugs)
            sheet.cell(row=row_number, column=8, value=unique_id)


        workbook.save("O:\\QA\\live.xlsx")
        messagebox.showinfo("Success", "Data added to Excel sheet.")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")

with open("site_config.json", "r") as config_file:
    site_config = json.load(config_file)

customer_label = tk.Label(data_frame, text="Customer:")
customer_label.grid(row=0, column=0, padx=5, pady=5, sticky="w")

customer_combo = ttk.Combobox(data_frame, values=["cemex", "Shafir", "hidenberg"])
customer_combo.grid(row=0, column=1, padx=5, pady=5, sticky="w")
customer_combo.bind("<<ComboboxSelected>>", update_site_options)

site_label = tk.Label(data_frame, text="Site(s):")
site_label.grid(row=1, column=0, padx=5, pady=5, sticky="w")

site_combo = ttk.Combobox(data_frame, values=[], state="readonly", justify="left", height=5)
site_combo.grid(row=1, column=1, padx=5, pady=5, sticky="w")
site_combo.set("Select Site(s)")
site_combo.bind("<<ComboboxSelected>>", update_net_version)

date_label = tk.Label(data_frame, text="Date:")
date_label.grid(row=2, column=0, padx=5, pady=5, sticky="w")

calendar = Calendar(data_frame)
calendar.grid(row=2, column=1, padx=5, pady=5, sticky="w")

net_label = tk.Label(data_frame, text="Net:")
net_label.grid(row=3, column=0, padx=5, pady=5, sticky="w")

net_entry = tk.Entry(data_frame)
net_entry.grid(row=3, column=1, padx=5, pady=5, sticky="w")

version_label = tk.Label(data_frame, text="Version:")
version_label.grid(row=4, column=0, padx=5, pady=5, sticky="w")

version_entry = tk.Entry(data_frame)
version_entry.grid(row=4, column=1, padx=5, pady=5, sticky="w")

total_event_label = tk.Label(data_frame, text="Total Event:")
total_event_label.grid(row=5, column=0, padx=5, pady=5, sticky="w")

total_event_entry = tk.Entry(data_frame)
total_event_entry.grid(row=5, column=1, padx=5, pady=5, sticky="w")

false_event_label = tk.Label(data_frame, text="False Event:")
false_event_label.grid(row=6, column=0, padx=5, pady=5, sticky="w")

false_event_entry = tk.Entry(data_frame)
false_event_entry.grid(row=6, column=1, padx=5, pady=5, sticky="w")

version_bugs_label = tk.Label(data_frame, text="Version Bugs:")
version_bugs_label.grid(row=7, column=0, padx=5, pady=5, sticky="w")

version_bugs_entry = tk.Entry(data_frame)
version_bugs_entry.grid(row=7, column=1, padx=5, pady=5, sticky="w")

unique_id_label = tk.Label(data_frame, text="Unique ID:")
unique_id_label.grid(row=8, column=0, padx=5, pady=5, sticky="w")

unique_id_entry = tk.Entry(data_frame)
unique_id_entry.grid(row=8, column=1, padx=5, pady=5, sticky="w")

add_data_button = tk.Button(data_frame, text="Add Data", command=add_data)
add_data_button.grid(row=9, column=0, columnspan=2, padx=5, pady=10)

# Clear Button
clear_button = tk.Button(data_frame, text="Clear", command=clear_fields)
clear_button.grid(row=9, column=2, padx=5, pady=10)

root.mainloop()


def open_website_and_login(website_url, username_value, password_value):
    # Create a new Chrome browser instance
    driver = webdriver.Chrome()

    # Open the specified website
    driver.get(website_url)

    # Use explicit wait to wait for the username field to be clickable
    wait = WebDriverWait(driver, 10)
    username_field = wait.until(EC.element_to_be_clickable((By.ID, "username")))

    # Click the username field to focus it
    username_field.click()

    # Send keys to the username field
    username_field.send_keys(username_value)

    # Similarly, handle the password field and login button

    password_field = wait.until(EC.element_to_be_clickable((By.ID, "password")))
    password_field.click()
    password_field.send_keys(password_value)

    login_button = wait.until(EC.element_to_be_clickable((By.ID, "kc-login")))
    login_button.click()

    # Wait for user confirmation before closing the browser
    input("Press Enter to close the browser...")
    driver.quit()  # Close the browser when the user presses Enter


# Create the main window
#root = tk.Tk()
#root.title("Website Automation")

# Create buttons
#button1 = tk.Button(root, text="CEMEX", command=lambda: open_website_and_login(
 #   "https://cemex-manager-app.ception.live/",
  #  "ceptionuser",
  #  "ceptioncemex?"
#))
#button2 = tk.Button(root, text="Shafir", command=lambda: open_website_and_login(
#    "https://shapir-manager-app.ception.live/",
#    "ceptionuser",
#   "ceptionshapir!"
#)
#button3 = tk.Button(root, text="Heidelberg", command=lambda: open_website_and_login(
#    "https://heidelberg-manager-app.ception.live/",
#    "ceptionuser",
#    "ceptionheidelberg%"
#))

# Pack buttons into the main window
#button1.pack()
#button2.pack()
#button3.pack()

# Start the GUI main loop
#root.mainloop()
